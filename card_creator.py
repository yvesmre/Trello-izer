from excel_parser import *
from variables import *
from trello_exports import *
from trello_imports import *
from datetime import date
from import_myob_data import *
import shutil
import tkinter
from tkinter import *
from openpyxl.styles import Font
import tkinter
from functools import partial
from trello_imports import *
import os, subprocess, platform
from openpyxl import load_workbook
from openpyxl.styles import Font
import shutil
from pandas import ExcelWriter
from openpyxl.cell import MergedCell
from variables import *

def all_children (wid) :
    _list = wid.winfo_children()

    for item in _list :
        if item.winfo_children() :
            _list.extend(item.winfo_children())

    return _list

def fit_text_to_widget(text_widget):
    # Get the number of lines and the longest line's length
    num_lines = int(text_widget.index('end-1c').split('.')[0])
    longest_line_length = max(len(line) for line in text_widget.get("1.0", "end-1c").split('\n'))

    # Calculate the widget's required height and width
    height = int(int(longest_line_length/(512/8))*2)
    width = int(min(longest_line_length, 512/8))

    # Resize the widget to fit the text
    text_widget.config(width=width, height=height)

def deploy_card(event):
    make_card(event.widget.get("1.0",'end-1c').split('-')[0])
    cards_to_be_made(event.widget.master)


def cards_to_be_made(screen):


    search_indicator = tkinter.Label(m, text="Searching...")
    search_indicator.config(bg="green")
    search_indicator.grid(row=2,column=1)


    def threadify_deploy_card(event):
        thread = Thread(target=deploy_card, args=(event, ))
        thread.start()

    import_myob()

    orders = parse_spreadsheets_for_orders(EXCEL_SPREADSHEET_READ)

    for child in all_children(screen):
        if(type(child) == tkinter.Text):
            child.destroy()

    for i in range(len(orders)):
        order = orders[i]

        myob = search_myob(order.job_number)
        if not myob:
            continue

        # lines = search_myob(order.job_number)["Lines"]
        # order.set_lines(lines)

        # if not (len(order.lines) > 0):
        #     continue

        display = tkinter.Text(screen)
        display.insert('1.0', str(order.job_number) + "-" + order.client)
        fit_text_to_widget(display)
        display.grid(row=2+i, column=1)
        display.bind("<Button-1>", threadify_deploy_card)

    print("Done looking for cards")
    search_indicator.destroy()


def create_spreadsheet(board_id, job_no, filename):

    #Hacky way of going about this, tbh
    if not os.path.exists(os.getcwd() + "/" + filename.split('/')[1]) :
        os.makedirs(os.getcwd() + "/" + filename.split('/')[1])

    board = import_cards_with_custom_fields_from_board(board_id)

    to_json = {
        "Item": [], 
               "Checklist": [], 
                'Part No.': []
                    }

    customer = ""
    index = 1

    user_to_name = {}
    desc = ""

    custom_fields = None
    for card in board:
        checklists = import_checklist(card['id'])

        card_name = card['name']
        desc = card['desc']
        if not (card_name.split('-')[0].replace('#', '').strip() == job_no): continue
        customer = card_name.split('-', 1)[1].strip()

        custom_fields = card['customFieldItems'] if 'customFieldItems' in card else None
        for obj in checklists:

            part_found = []
            for checklist_item in obj["checkItems"]:
                if("Part No." in checklist_item['name']):
                    to_json['Item'].append(int(index))
                    to_json['Checklist'].append(obj['name'])
                    to_json['Part No.'].append(checklist_item['name'].replace("Part No.", ''))
                    part_found.append(checklist_item['name'].replace("Part No.", ''))

            if not part_found: 

                to_json['Item'].append(int(index))
                to_json['Checklist'].append(obj['name'])
                to_json['Part No.'].append('N/A' if not part_found else ''.join([f'[{s}], ' if i < len(part_found) - 1 else f'[{s}]' for i, s in enumerate(part_found)]))
            
            index = index + 1
        break

    frame = pd.DataFrame(data=to_json)
    frame.index  = frame.index + 1

    shutil.copyfile(os.getcwd() + "/Build Sheet Trello Template.xlsx", os.getcwd() + filename)
    writer = ExcelWriter(os.getcwd() + filename, if_sheet_exists="overlay", mode='a', engine="openpyxl")
    

    frame.to_excel(excel_writer=writer,index=False, startrow=5, startcol=1, header=False)

    writer.close()

    wb = load_workbook(os.getcwd()+ filename)
    ws = wb.active

    manufacturer = 'N/A'
    model = 'N/A'
    vin = 'N/A'

    dealer = desc.split('Dealer:')[1].strip().split('Contact:')[0].strip()
    contact = desc.split('Contact:')[1].strip().split('#')[0].split('\n')[0].strip()

    if "##" in desc:
        desc = desc.split('##')[1].replace("**",'').strip()
        manufacturer = desc.split(' ')[0]
        model = desc.split(' ', 1)[1].split("#")[0]
        vin = '#' + desc.split('#')[1]
    elif(not custom_fields == None):
        for field in custom_fields:
            custom_field =import_custom_field(field['idCustomField'])
            
            if(custom_field['name'] == "MANUFACTURER"):
                manufacturer = import_custom_field_option(field['idCustomField'], field['idValue'])['value']['text']

            if(custom_field['name'] == "TRUCK MODEL"):
                model = field['value']['text']
            
            if(custom_field['name'] == 'CHASSIS VIN NO.'):
                vin = field['value']['text']

    ws.cell(row=2, column=3, value=manufacturer)
    ws.cell(row=3, column=3, value=model)
    ws.cell(row=4, column=3, value=vin)

    ws.cell(row=1, column=3, value=int(job_no)).font = Font(size = 27, bold=True)
    ws.cell(row=2, column=5, value=customer)
    ws.cell(row=3, column=5, value=dealer)
    ws.cell(row=4, column=5, value=contact)
    
    merge_start = 5

    for row in range(5, ws.max_row + 1):
        if ws.cell(row=row, column=3).value != ws.cell(row=row-1, column=3).value:
            if merge_start < row - 1:
                ws.merge_cells(start_row=merge_start, start_column=3, end_row=row-1, end_column=3)
            merge_start = row

    
    if merge_start < ws.max_row:
        ws.merge_cells(start_row=merge_start, start_column=3, end_row=ws.max_row, end_column=3)

        merge_start = 5
    
    merge_start_2 = 5
    for row in range(5, ws.max_row + 1):
        if ws.cell(row=row, column=2).value != ws.cell(row=row-1, column=2).value:
            if merge_start_2 < row - 1:
                ws.merge_cells(start_row=merge_start_2, start_column=2, end_row=row-1, end_column=2)
            merge_start_2 = row

    if merge_start_2 < ws.max_row:
        ws.merge_cells(start_row=merge_start_2, start_column=2, end_row=ws.max_row, end_column=2)

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            if cell.value:
                lines = int(len(str(cell.value)) / 20)
                ws.row_dimensions[cell.row].height = int(lines * 8) + 20

    wb.save(os.getcwd()+ filename)

    #Linux compatibility begone!
    # subprocess.call(['open', os.getcwd()+ filename]) if platform.system() == "Darwin" else os.startfile(os.getcwd()+ filename)



def make_card(job_no):

    import_myob()

    updates = {}
    orders = parse_spreadsheets_for_orders(EXCEL_SPREADSHEET_READ)

    if(len(orders) == 0):
        print("No cards to be created!")
        return
    
    order = None
    for ord in orders:
        if(str(ord.job_number) == job_no):
            order = ord
            break

    myob = search_myob(order.job_number)
    if myob:
        lines = search_myob(order.job_number)["Lines"]
        order.set_lines(lines)
    else: print(str(order.job_number) + " Does not have a MYOB entry! Skipping")
    
    if(len(order.lines) > 0) or True: # Only do job cards for ones with MYOB entries ### Could be a setting later

        fit_out_card = create_card(FIT_OUT_TODO_LIST if not USE_TESTING_LIST else TEST_LIST, str(order.job_number) + " - " + order.client, "", FIT_OUT_BUILD_TEMPLATE)
        fit_out_card_id = fit_out_card["id"]
        fit_out_card_url = fit_out_card["url"]
        
        checklists = import_checklist(fit_out_card_id)
        
        # Template Cards may come with Template Lists, disregard and delete. 
        # for checklist in checklists:
        #     delete_list(fit_out_card_id, checklist['id'])
        
        try:
            fit_out_description = "Dealer: "+ order.headers[1].split('-')[0] + "\nContact: " + order.headers[1].split('-')[1] + '\n'
        except:
            fit_out_description = ""

        if len(order.headers) > 2:
            for i in range(2, len(order.headers)):
                fit_out_description = fit_out_description + "\n##" + "**" + order.headers[i] + "**"
        update_card(fit_out_card_id, fit_out_description)
        drafting_card = create_card(DRAFTING_TODO_LIST if not USE_TESTING_LIST else TEST_LIST, str(order.job_number) + " - DRAFTING - " + order.client, "", DRAFTING_CARD_TEMPLATE)

        create_attachment(drafting_card["id"], fit_out_card_url)
        create_attachment(fit_out_card_id, drafting_card["url"])

        for line in order.lines:
            create_list(fit_out_card_id, line, BUILD_CHECKLIST_TEMPLATE)

        updates[str(order.job_number)] = date.today()

    if DIFFERENT_DESTINATION_EXCEL:
        shutil.copyfile(EXCEL_SPREADSHEET_READ, EXCEL_SPREADSHEET_WRITE)

    update_row(EXCEL_SPREADSHEET_WRITE if DIFFERENT_DESTINATION_EXCEL else EXCEL_SPREADSHEET_READ, updates) if WRITE_TO_EXCEL else None

    print("Run Done, made ", len(updates), " cards")

    # create_spreadsheet(board_id=TEST_BOARD if USE_TESTING_LIST else FIT_OUT_BOARD, job_no=job_no, filename= "/output/" + str(job_no) + ".xlsx")
        
        
if __name__ == "__main__":
    # warnings.simplefilter(action='ignore', category=FutureWarning)

    def look_for_cards():
        thread = Thread(target = cards_to_be_made, args=(m,))

        thread.start()


    
    main_window = tkinter.Tk()
    main_window.config(bg="black")
    
    # main_window.minsize(384, 384)
    # main_window.columnconfigure(1, weight=1)
    # main_window.rowconfigure(1, weight=1)

    main_frame = Frame(main_window)
    main_frame.pack(fill=BOTH, expand=1)

    main_canvas = Canvas(main_frame)
    main_canvas.pack(side=LEFT, fill=BOTH, expand=1)


    my_scrollbar = Scrollbar(main_frame, orient=VERTICAL, command=main_canvas.yview)
    my_scrollbar.pack(side=RIGHT, fill=Y)

    # configure the canvas
    main_canvas.configure(yscrollcommand=my_scrollbar.set)
    main_canvas.bind(
        '<Configure>', lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
    )

    m = Frame(main_canvas, width=384, height=384)

    # m.rowconfigure(list(range(2,50)), weight=1)

    start_button = tkinter.Button(m, text="Look for Cards to be Made", command=look_for_cards, width=24, height=4)
    start_button.grid(row=1, column=1)

    # start_button.place(x=30 , y=0)

    main_canvas.create_window((0, 0), window=m, anchor="nw")
    main_window.mainloop()

