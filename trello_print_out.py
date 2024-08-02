import tkinter
from functools import partial
from trello_imports import *
import os, subprocess, platform
from openpyxl import load_workbook
from openpyxl.styles import Font
import shutil
from pandas import ExcelWriter
from openpyxl.cell import MergedCell
from variables import *

def fit_text_to_widget(text_widget):
    # Get the number of lines and the longest line's length
    num_lines = int(text_widget.index('end-1c').split('.')[0])
    longest_line_length = max(len(line) for line in text_widget.get("1.0", "end-1c").split('\n'))

    # Calculate the widget's required height and width
    height = int(int(longest_line_length/(512/8))*2)
    width = int(min(longest_line_length, 512/8))

    # Resize the widget to fit the text
    text_widget.config(width=width, height=height)

def all_children (wid) :
    _list = wid.winfo_children()

    for item in _list :
        if item.winfo_children() :
            _list.extend(item.winfo_children())

    return _list


def create_spreadsheet(board_id, job_no, filename):

    #Hacky way of going about this, tbh
    if not os.path.exists(os.getcwd() + "/" + filename.split('/')[1]) :
        os.makedirs(os.getcwd() + "/" + filename.split('/')[1])

    search_indicator = tkinter.Label(m, text="Searching...")
    search_indicator.config(bg="green")
    search_indicator.grid(row=3,column=1)

    board = import_cards_with_custom_fields_from_board(board_id)

    to_json = {
        "Item": [], 
               "Checklist": [], 
                'Part No.': []
                    }




    customer = ""
    index = 1

    user_to_name = {}
    desc = ""

    custom_fields = None
    for card in board:
        checklists = import_checklist(card['id'])

        card_name = card['name']
        desc = card['desc']
        if not (card_name.split('-')[0].replace('#', '').strip() == job_no): continue
        customer = card_name.split('-', 1)[1].strip()

        custom_fields = card['customFieldItems'] if 'customFieldItems' in card else None
        for obj in checklists:

            part_found = []
            for checklist_item in obj["checkItems"]:
                if("Part No." in checklist_item['name']):
                    to_json['Item'].append(int(index))
                    to_json['Checklist'].append(obj['name'])
                    to_json['Part No.'].append(checklist_item['name'].replace("Part No.", ''))
                    part_found.append(checklist_item['name'].replace("Part No.", ''))

            if not part_found: 

                to_json['Item'].append(int(index))
                to_json['Checklist'].append(obj['name'])
                to_json['Part No.'].append('N/A' if not part_found else ''.join([f'[{s}], ' if i < len(part_found) - 1 else f'[{s}]' for i, s in enumerate(part_found)]))
            
            index = index + 1
        break

    frame = pd.DataFrame(data=to_json)
    frame.index  = frame.index + 1

    shutil.copyfile(os.getcwd() + "/Build Sheet Trello Template.xlsx", os.getcwd() + filename)
    writer = ExcelWriter(os.getcwd() + filename, if_sheet_exists="overlay", mode='a', engine="openpyxl")
    

    frame.to_excel(excel_writer=writer,index=False, startrow=5, startcol=1, header=False)

    writer.close()

    wb = load_workbook(os.getcwd()+ filename)
    ws = wb.active

    manufacturer = 'N/A'
    model = 'N/A'
    vin = 'N/A'

    dealer = desc.split('Dealer:')[1].strip().split('Contact:')[0].strip()
    contact = desc.split('Contact:')[1].strip().split('#')[0].split('\n')[0].strip()

    try: 
        if "##" in desc:
            desc = desc.split('##')[1].replace("**",'').strip()
            manufacturer = desc.split(' ')[0]
            model = desc.split(' ', 1)[1].split("#")[0]
            vin = '#' + desc.split('#')[1]
        elif(not custom_fields == None):
            for field in custom_fields:
                custom_field =import_custom_field(field['idCustomField'])
                
                if(custom_field['name'] == "MANUFACTURER"):
                    manufacturer = import_custom_field_option(field['idCustomField'], field['idValue'])['value']['text']

                if(custom_field['name'] == "TRUCK MODEL"):
                    model = field['value']['text']
                
                if(custom_field['name'] == 'CHASSIS VIN NO.'):
                    vin = field['value']['text']
    except: None

    
    ws.cell(row=2, column=3, value=manufacturer)
    ws.cell(row=3, column=3, value=model)
    ws.cell(row=4, column=3, value=vin)

    ws.cell(row=1, column=3, value=int(job_no)).font = Font(size = 27, bold=True)
    ws.cell(row=2, column=5, value=customer)
    ws.cell(row=3, column=5, value=dealer)
    ws.cell(row=4, column=5, value=contact)
    
    merge_start = 5

    for row in range(5, ws.max_row + 1):
        if ws.cell(row=row, column=3).value != ws.cell(row=row-1, column=3).value:
            if merge_start < row - 1:
                ws.merge_cells(start_row=merge_start, start_column=3, end_row=row-1, end_column=3)
            merge_start = row

    
    if merge_start < ws.max_row:
        ws.merge_cells(start_row=merge_start, start_column=3, end_row=ws.max_row, end_column=3)

        merge_start = 5
    
    merge_start_2 = 5
    for row in range(5, ws.max_row + 1):
        if ws.cell(row=row, column=2).value != ws.cell(row=row-1, column=2).value:
            if merge_start_2 < row - 1:
                ws.merge_cells(start_row=merge_start_2, start_column=2, end_row=row-1, end_column=2)
            merge_start_2 = row

    if merge_start_2 < ws.max_row:
        ws.merge_cells(start_row=merge_start_2, start_column=2, end_row=ws.max_row, end_column=2)

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            if cell.value:
                lines = int(len(str(cell.value)) / 20)
                ws.row_dimensions[cell.row].height = int(lines * 8) + 20

    wb.save(os.getcwd()+ filename)

    #Linux compatibility begone!
    subprocess.call(['open', os.getcwd()+ filename]) if platform.system() == "Darwin" else os.startfile(os.getcwd()+ filename)
    search_indicator.destroy()


if __name__ == "__main__":
    def search_card(widget):
        thread = Thread(target=create_spreadsheet, args=(FIT_OUT_BOARD if not USE_TESTING_LIST else TEST_BOARD, widget.get(), "/output/" + widget.get() + ".xlsx"))
        thread.start()
                 
    m = tkinter.Tk()
    m.config(bg="black")
    
    m.minsize(384, 384)
    m.columnconfigure(1, weight=1)

    tkinter.Label(m, text='Job No.').grid(row=0, column=1)

    e1 = tkinter.Entry(m)
    e1.grid(row=1, column=1)

    start_button = tkinter.Button(m, text="Get Card Details", command=partial(search_card, e1))
    start_button.grid(row=2, column=1)

    m.mainloop()

