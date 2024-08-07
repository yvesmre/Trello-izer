import openpyxl
import pandas as pd
import math
import datetime

keywords = ["Job No.", "Customer", "Client", "Quote Status", "Purchase Order Date", "Trello Card Created", "Invoice Date"]

class Order():
    def __init__(self, job_number, customer, client):
        self.job_number = job_number
        self.customer = customer
        self.client = client
        self.lines = []
        self.headers =[]
    
    def set_lines(self, lines):
        self.lines = []
        for line in lines:
            if(line["Type"] == "Header"): 
                self.headers.append(line["Description"])
            else:
                self.lines.append(line["Description"])

    def __str__(self):
        return f"{self.job_number} {self.customer} {self.client} {self.lines}"

def parse_spreadsheets_for_orders(file):
    data = []

    spreadsheet  = pd.read_excel(file, keep_default_na=False)

    column_index = None
    for index, row in spreadsheet.iterrows():
            if(row.iloc[0] == "Job No."):
               column_index = index

    spreadsheet.columns = spreadsheet.iloc[column_index]

    for index, row in spreadsheet.loc[:, keywords].iterrows():
        if (type(row["Job No."]) is int):
            if((row['Quote Status'] == "Won" or row['Quote Status'] == "PO Pending" or row['Quote Status'] == "Quoted") and type(row['Trello Card Created']) is not datetime.datetime):
                if(row['Trello Card Created'].lower() != "n/a" and row['Trello Card Created'].lower() != "new"):
                        if(not row['Trello Card Created']):
                            if(type(row["Invoice Date"]) is str):
                                order = Order(row['Job No.'], row['Customer'], row['Client'])
                                data.append(order)

    return data


def update_row(file, updates):
    updates = dict(updates)
    spreadsheet = openpyxl.load_workbook(file)

    sh = spreadsheet.active 
    trello_card_created_column = 19
    for row in sh.iter_rows():
        if(row[0].value == "Job No."):
            for i in range(len(row)):
                if(row[i].value == "Trello Card Created"):
                    trello_card_created_column = i
                    break

    for row in sh.iter_rows(): 
        if str(row[0].value) in updates:
            row[trello_card_created_column].value = updates.pop(str(row[0].value))

    spreadsheet.save(file)
        

    